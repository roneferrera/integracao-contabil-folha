# ============================================================
# app.py  –  Integração Contábil Domínio V6.0
# Classificação: Evento vs Conta Contábil por Grupo de Despesa
# Tipo Folha "Empresa" = INSS Patronal com mapeamento específico
# Sistema de pontuação positivo/negativo para contas de folha
# Regras específicas por rubrica: Desconto, Adiantamento, Rescisão
# ============================================================

import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERSAO = "V6.0"

# ══════════════════════════════════════════════════════════════════════════
# TEMA
# ══════════════════════════════════════════════════════════════════════════
def apply_tr_theme():
    st.markdown("""
        <style>
        html, body, [class*="css"] {
            font-family: 'Segoe UI', 'Arial', sans-serif; color: #444444;
        }
        h1, h2, h3 { color: #FF8000; font-weight: 700; }
        section[data-testid="stSidebar"] { background-color: #444444; }
        section[data-testid="stSidebar"] * { color: #FFFFFF !important; }
        .stButton > button {
            background-color: #FF8000; color: #FFFFFF;
            border: none; border-radius: 4px; font-weight: bold;
        }
        .stButton > button:hover { background-color: #D64001; }
        .stDownloadButton > button {
            background-color: #FF8000; color: #FFFFFF;
            border: none; border-radius: 4px; font-weight: bold;
        }
        .stDownloadButton > button:hover { background-color: #D64001; }
        div[data-testid="stExpander"] { border: 1px solid #FF8000; border-radius: 6px; }
        </style>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
# NORMALIZAÇÃO
# ══════════════════════════════════════════════════════════════════════════
def _norm(texto: str) -> str:
    t = texto.upper()
    for orig, sub in [
        ("Ã","A"),("Á","A"),("Â","A"),("À","A"),("Ä","A"),
        ("É","E"),("Ê","E"),("È","E"),("Ë","E"),
        ("Í","I"),("Î","I"),("Ï","I"),
        ("Ó","O"),("Ô","O"),("Õ","O"),("Ö","O"),
        ("Ú","U"),("Û","U"),("Ü","U"),
        ("Ç","C"),("Ñ","N"),
    ]:
        t = t.replace(orig, sub)
    return t


# ══════════════════════════════════════════════════════════════════════════
# SISTEMA DE PONTUAÇÃO
# ══════════════════════════════════════════════════════════════════════════
PALAVRAS_FOLHA_POSITIVO: list[tuple[str, int]] = [
    ("SALARIOS E ORDENADOS A PAGAR", 10), ("SALARIOS A PAGAR", 10),
    ("PRO-LABORE A PAGAR", 10), ("GRATIFICACOES A PAGAR", 10),
    ("FERIAS A PAGAR", 10), ("RESCISOES A PAGAR", 10),
    ("13 SALARIO A PAGAR", 10), ("PENSAO ALIMENTICIA A PAGAR", 10),
    ("COMISSOES A PAGAR", 10), ("AUTONOMOS A PAGAR", 10),
    ("INDENIZACOES A PAGAR", 10), ("PREMIOS E BONIFICACOES", 10),
    ("INSS A RECOLHER", 10), ("INSS SOBRE PROVISOES", 10),
    ("FGTS A RECOLHER", 10), ("FGTS SOBRE PROVISOES", 10),
    ("PIS S/ FOLHA A RECOLHER", 10), ("IRRF S/ FOLHA", 10),
    ("CONTRIBUICOES SINDICAIS", 10),
    ("PROVISOES PARA FERIAS", 10), ("PROVISOES PARA 13", 10),
    ("INSS SOBRE PROVISOES PARA FERIAS", 10), ("INSS SOBRE PROVISOES PARA 13", 10),
    ("FGTS SOBRE PROVISOES PARA FERIAS", 10), ("FGTS SOBRE PROVISOES PARA 13", 10),
    ("OBRIGACOES COM O PESSOAL", 10), ("OBRIGACOES SOCIAIS", 10),
    ("OBRIGACOES TRABALHISTA", 10), ("OBRIGACOES TRABALHISTAS E PREVIDENCIARIA", 10),
    ("PROVISOES", 8),
    ("DESPESAS COM PESSOAL", 10), ("SALARIOS E ORDENADOS", 8),
    ("PRO-LABORE", 8), ("PREMIOS E GRATIFICACOES", 8),
    ("INDENIZACOES E AVISO PREVIO", 8), ("ASSISTENCIA MEDICA E SOCIAL", 8),
    ("VALE TRANSPORTE", 7), ("VALE REFEICAO", 7),
    ("ALIMENTACAO/ CESTA BASICA", 7), ("DESPESAS COM ALIMENTACAO", 7),
    ("PENSAO ALIMENTICIA", 7), ("COMISSOES SOBRE VENDAS", 7),
    ("COMISSOES", 7), ("HORAS EXTRAS", 7), ("PIS S/ FOLHA", 7),
    ("MAO-DE-OBRA DIRETA", 10), ("MAO-DE-OBRA INDIRETA", 10),
    ("SALARIOS E ORDENADOS CUSTOS", 10), ("PRO-LABORE CUSTOS", 10),
    ("FERIAS CUSTOS", 10), ("INSS CUSTOS", 10), ("FGTS CUSTOS", 10),
    ("INSS EMPRESA", 10), ("INSS TERCEIROS", 10), ("INSS ACIDENTE", 10),
    ("INSS PATRONAL", 10), ("ENCARGOS SOCIAIS", 10),
    ("CONTRIBUICAO PREVIDENCIARIA", 10), ("CONTRIBUICAO PATRONAL", 10),
    ("IMPOSTO DE RENDA A RECOLHER", 6), ("IMPOSTO DE RENDA RETIDO", 6),
    ("IRRF", 6), ("INSS", 5), ("FGTS", 5),
    ("FOLHA DE PAGAMENTO", 8), ("REMUNERACAO", 6), ("PESSOAL", 5),
    ("TRABALHISTA", 5), ("PREVIDENCIARIA", 5), ("PATRONAL", 8),
    ("EMPREGADOS", 5), ("FUNCIONARIOS", 5),
    # Adiantamentos (Ativo)
    ("ADIANTAMENTO DE SALARIO", 10), ("ADIANTAMENTO DE 13", 10),
    ("ADIANTAMENTOS DE FERIAS", 10), ("ADIANTAMENTO A EMPREGADOS", 8),
    ("ADIANTAMENTOS", 5),
]

PALAVRAS_FOLHA_NEGATIVO: list[tuple[str, int]] = [
    ("FORNECEDORES NACIONAIS", -50), ("FORNECEDORES ESTRANGEIROS", -50),
    ("FORNECEDORES DO GRUPO", -50), ("FORNECEDORES", -40),
    ("CLIENTES NACIONAIS", -50), ("CLIENTES ESTRANGEIROS", -50),
    ("CLIENTES RELACIONADOS", -50), ("CLIENTES", -40),
    ("MERCADORIAS PARA REVENDA", -50), ("MATERIA-PRIMA", -30),
    ("ESTOQUE", -40), ("ALMOXARIFADO", -30),
    ("PRODUTOS ACABADOS", -40), ("PRODUTOS SEMI ACABADOS", -40),
    ("IMOVEIS", -40), ("MAQUINAS E EQUIPAMENTOS", -40),
    ("VEICULOS", -40), ("MOVEIS E UTENSILIOS", -40),
    ("COMPUTADORES E ACESSORIOS", -40), ("INSTALACOES", -40),
    ("FERRAMENTAS E ACESSORIOS", -40), ("SOFTWARES", -40),
    ("MARCAS E PATENTES", -40), ("DEPRECIACAO DE EDIFICIOS", -30),
    ("DEPRECIACAO DE MOVEIS", -30), ("DEPRECIACAO DE MAQUINAS", -30),
    ("DEPRECIACAO DE VEICULOS", -30),
    ("BANCO DO BRASIL", -30), ("BANCO ITAU", -30), ("BANCO BRADESCO", -30),
    ("BANCO SANTANDER", -30), ("BANCO INTER", -30), ("BANCO C6", -30),
    ("BANCO NU", -30), ("BANCO CORA", -30), ("BANCO DAYCOVAL", -30),
    ("CAIXA ECONOMICA", -30), ("CAIXA GERAL", -30),
    ("FUNDO FIXO DE CAIXA", -30), ("APLICACOES FINANCEIRAS", -30),
    ("APLICACOES BANCO", -30), ("CHEQUE ESPECIAL", -30),
    ("EMPRESTIMOS BANCOS", -30), ("FINANCIAMENTO BANCO", -30),
    ("IPI A RECOLHER", -30), ("ICMS A RECOLHER", -30),
    ("ISS A RECOLHER", -30), ("PIS A RECOLHER", -30),
    ("COFINS A RECOLHER", -30), ("SIMPLES NACIONAL A RECOLHER", -30),
    ("VENDA DE PRODUTOS", -50), ("VENDA DE MERCADORIAS", -50),
    ("SERVICOS PRESTADOS", -40), ("RECEITA", -40),
    ("CAPITAL SOCIAL", -50), ("RESERVAS", -40),
    ("LUCROS OU PREJUIZOS", -40), ("DIVIDENDOS", -40),
    ("RESULTADO DO EXERCICIO", -40), ("APURACAO DO RESULTADO", -40),
    ("ADIANTAMENTO A SOCIOS", -30), ("ADIANTAMENTO A FORNECEDORES", -20),
    ("TITULOS A RECEBER", -30), ("DEPOSITOS JUDICIAIS", -30),
    ("INVESTIMENTOS", -30), ("PARTICIPACOES SOCIETARIAS", -30),
    ("IPI A RECUPERAR", -30), ("ICMS A RECUPERAR", -30),
    ("PIS A RECUPERAR", -30), ("COFINS A RECUPERAR", -30),
]

SCORE_MINIMO_FOLHA = 5


def calcular_score_folha(nome_conta_norm: str) -> int:
    score = 0
    for termo, peso in PALAVRAS_FOLHA_POSITIVO:
        if _norm(termo) in nome_conta_norm:
            score += peso
    for termo, peso in PALAVRAS_FOLHA_NEGATIVO:
        if _norm(termo) in nome_conta_norm:
            score += peso
    return score


def conta_e_de_folha(nome_conta_norm: str) -> bool:
    return calcular_score_folha(nome_conta_norm) >= SCORE_MINIMO_FOLHA


# ══════════════════════════════════════════════════════════════════════════
# CLASSIFICADOR LOCAL DE RUBRICAS
# ══════════════════════════════════════════════════════════════════════════
KWORDS_RUBRICA: dict[str, list[str]] = {
    "Custo Direto de Produção": [
        "MATERIA PRIMA","MATERIAL APLICADO","MAO DE OBRA DIRETA",
        "SALARIOS E ORDENADOS CUSTOS","PRO LABORE CUSTOS",
        "13 SALARIO CUSTOS","FERIAS CUSTOS","INSS CUSTOS",
        "FGTS CUSTOS","INDUSTRIALIZACAO","PRODUCAO",
    ],
    "Custo Direto de Serviços": [
        "SERVICOS PRESTADOS","MAO DE OBRA","PRESTACAO DE SERVICO",
        "TERCEIRIZACAO","SUBCONTRATACAO","CUSTOS SERVICOS",
    ],
    "Custo Indireto de Produção": [
        "OVERHEAD","MANUTENCAO","DEPRECIACAO","ENERGIA ELETRICA",
        "AGUA ESGOTO","COMBUSTIVEL","ALUGUEL EQUIPAMENTO",
        "MATERIAL CONSUMO INDIRETO","LOCACAO","CONDOMINIO",
    ],
    "Despesa com Vendas": [
        "COMISSAO","COMISSOES","PROPAGANDA","PUBLICIDADE",
        "BONIFICACAO","REPRESENTACAO","FRETE VENDA",
    ],
    "Despesa Financeira": [
        "JUROS","IOF","TARIFA BANCARIA","VARIACAO CAMBIAL",
        "VARIACAO MONETARIA","DESCONTO CONCEDIDO","MORA",
        "EMPRESTIMO","FINANCIAMENTO",
    ],
    "Despesa Não Operacional": [
        "IRPJ","CSLL","PROVISAO IR","PROVISAO CSLL",
        "PERDA","ALIENACAO","SINISTRO","BAIXA ATIVO",
    ],
    "Despesa Administrativa": [
        "SALARIO","ORDENADO","PRO LABORE","GRATIFICACAO",
        "BONUS","PREMIO","HORAS EXTRAS","ADICIONAL NOTURNO",
        "ADICIONAL","INSALUBRIDADE","PERICULOSIDADE",
        "FERIAS","ABONO","13 SALARIO","AVISO PREVIO",
        "RESCISAO","INDENIZACAO","LICENCA","AFASTAMENTO",
        "REPOUSO","DSR","REFLEXO","MEDIA","DIFERENCA",
        "ADIANTAMENTO","VALE TRANSPORTE","VALE REFEICAO",
        "ALIMENTACAO","CESTA BASICA","AUXILIO","REEMBOLSO",
        "PARTICIPACAO LUCROS","PLR","COMPL SALARIAL",
        "DESC ADIANT","DESCONTO ADIANT","INSS","FGTS",
        "IMPOSTO DE RENDA","IRRF","PENSAO ALIMENTICIA",
        "PLANO SAUDE","PLANO ODONTOLOGICO","COPARTICIPACAO",
        "EMPRESTIMO CONSIGNADO","SINDICATO","CONTRIBUICAO",
        "DIAS NORMAIS","HORAS NORMAIS","SALDO SALARIO",
        "ESTOURO","TROCO","FECHAMENTO","DIAS FERIAS",
    ],
    "Encargo Patronal": [
        "INSS EMPRESA","INSS TERCEIROS","INSS ACID","INSS ACIDENTE",
        "INSS PATRONAL","ENCARGO PATRONAL","CONTRIBUICAO PATRONAL",
        "CONTRIBUICAO PREVIDENCIARIA PATRONAL",
        "RAT","FAP","SISTEMA S","SESI","SENAI","SEBRAE","SESC","SENAC",
        "INSS 13","INSS FERIAS","INSS EMPRESA FERIAS","INSS EMPRESA 13",
    ],
}

TIPOS_NAO_CUSTO = {"Desconto", "Informativa", "Inf. Dedutora"}


def classificar_rubrica_local(nome_rubrica: str, tipo_rubrica: str, tipo_folha: str = "1") -> dict:
    if tipo_folha == "2":
        return {"grupo": "Encargo Patronal", "confianca": "alta"}

    nome_norm = _norm(nome_rubrica)
    grupos_candidatos = list(KWORDS_RUBRICA.keys())

    if tipo_rubrica in TIPOS_NAO_CUSTO:
        grupos_candidatos = [g for g in grupos_candidatos if "Custo" not in g]

    scores: dict[str, int] = {g: 0 for g in grupos_candidatos}
    for grupo in grupos_candidatos:
        for kw in KWORDS_RUBRICA[grupo]:
            kw_norm = _norm(kw)
            if kw_norm in nome_norm:
                scores[grupo] += len(kw_norm.split())

    melhor_grupo = max(scores, key=lambda g: scores[g])
    melhor_score = scores[melhor_grupo]

    if melhor_score >= 4:   confianca = "alta"
    elif melhor_score >= 2: confianca = "media"
    elif melhor_score >= 1: confianca = "baixa"
    else:
        melhor_grupo = "Despesa Administrativa"
        confianca    = "baixa"

    return {"grupo": melhor_grupo, "confianca": confianca}


# ══════════════════════════════════════════════════════════════════════════
# PARSE DO PLANO DE CONTAS
# ══════════════════════════════════════════════════════════════════════════
def parse_plano_contas(file_bytes: bytes, filename: str, log: list) -> pd.DataFrame:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else "xlsx"
    df_raw = None

    if ext == "xlsx":
        try:
            df_raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=0, dtype=str, engine="openpyxl")
            log.append("Plano de Contas: lido como .xlsx.")
        except Exception as e:
            log.append(f"ERRO ao abrir .xlsx: {e}")
            return pd.DataFrame()
    else:
        for engine in ["xlrd", "openpyxl"]:
            try:
                df_raw = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=0, dtype=str, engine=engine)
                log.append(f"Plano de Contas: lido como .xls (engine={engine}).")
                break
            except Exception as e:
                log.append(f"  engine={engine} falhou: {e}")
                df_raw = None

    if df_raw is None:
        log.append("ERRO: Não foi possível abrir o Plano de Contas.")
        return pd.DataFrame()

    cols = [str(c).strip() for c in df_raw.columns]
    idx_empresa = idx_reduzido = idx_classificacao = idx_tipo = idx_descricao = None

    for i, c in enumerate(cols):
        cl = c.lower()
        if idx_empresa is None and (i == 0 or "plano de contas" in cl or (cl == "empresa" and i < 3)):
            idx_empresa = i
        if idx_reduzido is None and ("reduzido" in cl or "unnamed: 1" in cl):
            idx_reduzido = i
        if idx_classificacao is None and ("classifica" in cl or "unnamed: 2" in cl):
            idx_classificacao = i
        if idx_tipo is None and ("unnamed: 3" in cl or cl == "tipo" or ("tipo" in cl and "ecf" not in cl and i < 6)):
            idx_tipo = i
        if idx_descricao is None and ("descri" in cl or "unnamed: 4" in cl):
            idx_descricao = i

    if idx_empresa is None:       idx_empresa = 0
    if idx_reduzido is None:      idx_reduzido = 1
    if idx_classificacao is None: idx_classificacao = 2
    if idx_tipo is None:          idx_tipo = 3
    if idx_descricao is None:     idx_descricao = 4

    registros = []
    ignorados = 0

    for _, row in df_raw.iterrows():
        empresa_val = str(row.iloc[idx_empresa]).strip()
        reduzido    = str(row.iloc[idx_reduzido]).strip()
        classif     = str(row.iloc[idx_classificacao]).strip()
        tipo_raw    = str(row.iloc[idx_tipo]).strip().upper()
        nome        = str(row.iloc[idx_descricao]).strip()

        if empresa_val.lower().startswith("total"): ignorados += 1; continue
        if empresa_val.lower() in ("nan","none",""): ignorados += 1; continue
        if classif.endswith(".0"): classif = classif[:-2]
        if not re.match(r'^\d+$', classif): ignorados += 1; continue
        if tipo_raw not in ("S","A"): ignorados += 1; continue
        if not nome or nome.lower() in ("nan","none",""): ignorados += 1; continue

        if reduzido.endswith(".0"):
            reduzido = reduzido[:-2]
        if reduzido.lower() in ("nan","none",""):
            reduzido = classif

        nome_norm   = _norm(nome)
        score_folha = calcular_score_folha(nome_norm)

        registros.append({
            "reduzido":      reduzido,
            "classificacao": classif,
            "nome_conta":    nome_norm,
            "nome_original": nome,
            "tipo":          tipo_raw,
            "score_folha":   score_folha,
        })

    df = pd.DataFrame(registros).drop_duplicates(subset=["classificacao"]).reset_index(drop=True)
    n_a     = len(df[df["tipo"] == "A"])
    n_s     = len(df[df["tipo"] == "S"])
    n_folha = len(df[(df["tipo"] == "A") & (df["score_folha"] >= SCORE_MINIMO_FOLHA)])
    log.append(
        f"Plano de Contas OK: {len(df)} contas ({n_a} analíticas · {n_s} sintéticas · "
        f"{ignorados} ignoradas · {n_folha} analíticas de folha)"
    )
    return df


# ══════════════════════════════════════════════════════════════════════════
# KEYWORDS POR GRUPO — DÉBITO E CRÉDITO
# ══════════════════════════════════════════════════════════════════════════
KWORDS_DEBITO: dict[str, list[str]] = {
    "Custo Direto de Produção": [
        "MATERIA-PRIMA","MATERIAL APLICADO","MAO-DE-OBRA DIRETA",
        "SALARIOS E ORDENADOS CUSTOS","PRO-LABORE CUSTOS",
        "13 SALARIO CUSTOS","FERIAS CUSTOS","INSS CUSTOS","FGTS CUSTOS",
        "INDUSTRIALIZACAO","CUSTOS DIRETOS DE PRODUCAO",
    ],
    "Custo Direto de Serviços": [
        "CUSTOS DIRETOS DA PRODUCAO DE SERVICOS","MAO-DE-OBRA DIRETA",
        "SALARIOS E ORDENADOS","INSS","FGTS","FERIAS","13 SALARIO",
        "CUSTOS SERVICOS","VALE TRANSPORTE","ALIMENTACAO",
    ],
    "Custo Indireto de Produção": [
        "MAO-DE-OBRA INDIRETA","MATERIAIS DE CONSUMO INDIRETO",
        "ALUGUEIS","DEPRECIACOES","COMBUSTIVEIS","ENERGIA ELETRICA",
        "AGUA E ESGOTO","CUSTOS INDIRETOS","MANUTENCAO","LOCACAO",
    ],
    "Despesa Administrativa": [
        "DESPESAS COM PESSOAL","SALARIOS E ORDENADOS","PRO-LABORE",
        "PREMIOS E GRATIFICACOES","13 SALARIO","FERIAS","INSS","FGTS",
        "INDENIZACOES","ASSISTENCIA MEDICA","VALE TRANSPORTE",
        "PIS S/ FOLHA","ALIMENTACAO","VALE REFEICAO","HORAS EXTRAS",
        "PENSAO ALIMENTICIA","CESTA BASICA","COMISSOES",
        "ALUGUEIS","ENERGIA ELETRICA","AGUA E ESGOTO","TELEFONE",
        "SEGUROS","MATERIAL DE ESCRITORIO","DEPRECIACAO",
        "COMBUSTIVEIS","MATERIAIS DE CONSUMO","CONDOMINIOS",
        "DESPESAS GERAIS","FRETES E CARRETOS","MANUTENCAO",
        "VIAGENS","REFEICOES","SERVICOS TOMADOS","BENS DE PEQUENO VALOR",
    ],
    "Despesa com Vendas": [
        "DESPESAS COM VENDAS","COMISSOES SOBRE VENDAS","COMISSOES",
        "PROPAGANDA E PUBLICIDADE","BONIFICACAO","FRETES E CARRETOS",
        "MANUTENCAO DE VEICULOS","VIAGENS","ALUGUEIS",
    ],
    "Despesa Financeira": [
        "DESPESAS FINANCEIRAS","JUROS PASSIVOS","VARIACOES MONETARIAS",
        "VARIACOES CAMBIAIS","DESCONTOS FINANCEIROS CONCEDIDOS",
        "JUROS DE MORA","JUROS E COMISSOES BANCARIAS",
        "JUROS SOBRE EMPRESTIMOS","MULTAS PASSIVAS","TARIFA BANCARIA",
        "EMPRESTIMO / FINANCIAMENTO","IOF",
    ],
    "Despesa Não Operacional": [
        "DESPESAS NAO OPERACIONAIS","PERDAS NA ALIENACAO",
        "RESULTADO NEGATIVO NA ALIENACAO","RESULTADO NEGATIVO DE SINISTRO",
        "OUTRAS BAIXAS DO ATIVO","BAIXAS DE INVESTIMENTOS",
        "BAIXAS DE IMOBILIZADO","PROVISAO IRPJ","PROVISAO CSLL",
        "PERDAS POR FALTA NO INVENTARIO",
    ],
    "Encargo Patronal": [
        "DESPESAS COM PESSOAL","INSS","ENCARGOS SOCIAIS",
        "CONTRIBUICAO PREVIDENCIARIA","CONTRIBUICAO PATRONAL",
        "SALARIOS E ORDENADOS","CUSTOS DIRETOS",
        "MAO-DE-OBRA DIRETA","MAO-DE-OBRA INDIRETA",
        "INSS CUSTOS","FGTS CUSTOS",
    ],
}

KWORDS_CREDITO: dict[str, list[str]] = {
    "Custo Direto de Produção": [
        "SALARIOS E ORDENADOS A PAGAR","PRO-LABORE A PAGAR",
        "FERIAS A PAGAR","13 SALARIO A PAGAR","INSS A RECOLHER",
        "FGTS A RECOLHER","PROVISOES PARA FERIAS","PROVISOES PARA 13",
        "OBRIGACOES COM O PESSOAL","OBRIGACOES SOCIAIS","PROVISOES",
        "OBRIGACOES TRABALHISTA",
    ],
    "Custo Direto de Serviços": [
        "SALARIOS E ORDENADOS A PAGAR","FERIAS A PAGAR","13 SALARIO A PAGAR",
        "INSS A RECOLHER","FGTS A RECOLHER","PROVISOES PARA FERIAS",
        "PROVISOES PARA 13","OBRIGACOES COM O PESSOAL","OBRIGACOES SOCIAIS",
        "PROVISOES","OBRIGACOES TRABALHISTA",
    ],
    "Custo Indireto de Produção": [
        "SALARIOS E ORDENADOS A PAGAR","FERIAS A PAGAR","INSS A RECOLHER",
        "FGTS A RECOLHER","PROVISOES PARA FERIAS","PROVISOES PARA 13",
        "OBRIGACOES COM O PESSOAL","OBRIGACOES SOCIAIS","PROVISOES",
        "CONTAS A PAGAR","ALUGUEIS A PAGAR",
    ],
    "Despesa Administrativa": [
        "SALARIOS E ORDENADOS A PAGAR","PRO-LABORE A PAGAR",
        "GRATIFICACOES A PAGAR","FERIAS A PAGAR","RESCISOES A PAGAR",
        "13 SALARIO A PAGAR","PENSAO ALIMENTICIA A PAGAR",
        "COMISSOES A PAGAR","AUTONOMOS A PAGAR","INDENIZACOES A PAGAR",
        "INSS A RECOLHER","FGTS A RECOLHER","PIS S/ FOLHA A RECOLHER",
        "IRRF S/ FOLHA","CONTRIBUICOES SINDICAIS",
        "PROVISOES PARA FERIAS","PROVISOES PARA 13",
        "INSS SOBRE PROVISOES","FGTS SOBRE PROVISOES",
        "OBRIGACOES COM O PESSOAL","OBRIGACOES SOCIAIS","PROVISOES",
        "OBRIGACOES TRABALHISTA","CONTAS A PAGAR",
        "ENERGIA ELETRICA A PAGAR","TELEFONE A PAGAR",
        "ALUGUEIS A PAGAR","OUTRAS OBRIGACOES",
    ],
    "Despesa com Vendas": [
        "SALARIOS E ORDENADOS A PAGAR","FERIAS A PAGAR","13 SALARIO A PAGAR",
        "INSS A RECOLHER","FGTS A RECOLHER","PROVISOES PARA FERIAS",
        "PROVISOES PARA 13","OBRIGACOES COM O PESSOAL","OBRIGACOES SOCIAIS",
        "PROVISOES","OBRIGACOES TRABALHISTA","CONTAS A PAGAR","OUTRAS OBRIGACOES",
    ],
    "Despesa Financeira": [
        "CONTAS A PAGAR","OUTRAS OBRIGACOES",
        "IMPOSTOS E CONTRIBUICOES A RECOLHER",
    ],
    "Despesa Não Operacional": [
        "CONTAS A PAGAR","OUTRAS OBRIGACOES",
        "IMPOSTOS E CONTRIBUICOES A RECOLHER",
        "PROVISAO PARA IMPOSTO DE RENDA S/ LUCRO",
        "PROVISAO P/ CONTRIBUICAO SOCIAL S/ LUCRO",
        "IMPOSTO DE RENDA A RECOLHER","CONTRIBUICAO SOCIAL A RECOLHER",
    ],
    "Encargo Patronal": [
        "INSS A RECOLHER","OBRIGACOES SOCIAIS",
        "OBRIGACOES TRABALHISTA","OBRIGACOES TRABALHISTAS E PREVIDENCIARIA",
        "IMPOSTOS E CONTRIBUICOES A RECOLHER",
        "CONTRIBUICOES SINDICAIS","PIS S/ FOLHA A RECOLHER",
    ],
}

GRUPOS_LISTA = [
    "Despesa Administrativa", "Despesa com Vendas", "Despesa Financeira",
    "Despesa Não Operacional", "Custo Direto de Produção",
    "Custo Direto de Serviços", "Custo Indireto de Produção",
    "Encargo Patronal", "Outro",
]


# ══════════════════════════════════════════════════════════════════════════
# FUNÇÕES DE CLASSIFICAÇÃO DE CONTAS
# ══════════════════════════════════════════════════════════════════════════
def _conta_bate(nome_conta_norm: str, keywords: list[str]) -> bool:
    for kw in keywords:
        if _norm(kw) in nome_conta_norm:
            return True
    return False


def _analiticas(df: pd.DataFrame) -> pd.DataFrame:
    return df[df["tipo"] == "A"].copy() if not df.empty else df


def _analiticas_folha(df: pd.DataFrame) -> pd.DataFrame:
    df_a = _analiticas(df)
    if df_a.empty:
        return df_a
    if "score_folha" in df_a.columns:
        filtrado = df_a[df_a["score_folha"] >= SCORE_MINIMO_FOLHA]
        return filtrado if not filtrado.empty else df_a
    return df_a


def _fmt_opcoes(df_f: pd.DataFrame) -> list[str]:
    return [""] + [
        f"{r['reduzido']} - {r['nome_original']}"
        for _, r in df_f.iterrows()
    ]


def filtrar_contas_por_grupo(
    df_contas: pd.DataFrame,
    grupo: str,
    aplicar_filtro_folha: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if aplicar_filtro_folha and grupo != "Outro":
        df_a = _analiticas_folha(df_contas)
    else:
        df_a = _analiticas(df_contas)

    if df_a.empty:
        return pd.DataFrame(), pd.DataFrame()

    kw_d = KWORDS_DEBITO.get(grupo, [])
    kw_c = KWORDS_CREDITO.get(grupo, [])

    if kw_d and grupo != "Outro":
        mask_d = df_a["nome_conta"].apply(lambda n: _conta_bate(n, kw_d))
        df_d = df_a[mask_d]
    else:
        df_d = df_a

    if kw_c and grupo != "Outro":
        mask_c = df_a["nome_conta"].apply(lambda n: _conta_bate(n, kw_c))
        df_c = df_a[mask_c]
    else:
        df_c = df_a

    return df_d, df_c


def classificar_contas(df_contas: pd.DataFrame, grupo: str) -> tuple[list[str], list[str]]:
    df_d, df_c = filtrar_contas_por_grupo(df_contas, grupo)
    return _fmt_opcoes(df_d), _fmt_opcoes(df_c)


def extrair_codigo(opcao: str) -> str:
    if not opcao or " - " not in opcao:
        return opcao or ""
    return opcao.split(" - ")[0].strip()


def extrair_descricao(opcao: str) -> str:
    if not opcao or " - " not in opcao:
        return ""
    partes = opcao.split(" - ", 1)
    return partes[1].strip() if len(partes) > 1 else ""


def buscar_conta_por_reduzido(df_contas: pd.DataFrame, reduzido: str) -> str:
    if df_contas is None or df_contas.empty or not reduzido:
        return ""
    mask = df_contas["reduzido"] == str(reduzido).strip()
    resultado = df_contas[mask]
    return resultado.iloc[0]["nome_original"] if not resultado.empty else ""


def _idx(opcoes: list[str], valor: str) -> int:
    if not valor:
        return 0
    for i, op in enumerate(opcoes):
        if op.startswith(valor):
            return i
    return 0


# ══════════════════════════════════════════════════════════════════════════
# FUNÇÃO AUXILIAR: MELHOR CONTA POR PONTUAÇÃO
# ══════════════════════════════════════════════════════════════════════════
def _melhor_conta_por_score(
    df_contas: pd.DataFrame,
    termos_pos: list[tuple[str, int]],
    termos_neg: list[tuple[str, int]],
    score_min: int = 1,
) -> tuple[str, str]:
    """
    Percorre todas as contas analíticas e pontua cada uma.
    Retorna (reduzido, nome_original) da conta com maior score >= score_min.
    """
    df_a = _analiticas(df_contas)
    if df_a.empty:
        return "", ""

    melhor_score = score_min - 1
    melhor_red   = ""
    melhor_nome  = ""

    for _, row in df_a.iterrows():
        nome = row["nome_conta"]  # já normalizado
        score = 0
        for termo, peso in termos_pos:
            if _norm(termo) in nome:
                score += peso
        for termo, peso in termos_neg:
            if _norm(termo) in nome:
                score += peso  # peso negativo

        if score > melhor_score:
            melhor_score = score
            melhor_red   = row["reduzido"]
            melhor_nome  = row["nome_original"]

    return melhor_red, melhor_nome


# ══════════════════════════════════════════════════════════════════════════
# TABELAS DE PONTUAÇÃO ESPECÍFICAS POR RUBRICA
# ══════════════════════════════════════════════════════════════════════════

# ── Débito padrão para descontos: Salários e Ordenados a Pagar (187) ──────
DESCONTO_DEBITO_SAL_PAGAR_POS = [
    ("SALARIOS E ORDENADOS A PAGAR", 500),
    ("SALARIOS A PAGAR", 450),
    ("OBRIGACOES COM O PESSOAL", 300),
    ("GRATIFICACOES A PAGAR", 200),
]
DESCONTO_DEBITO_SAL_PAGAR_NEG = [
    ("DESPESA", -200), ("CUSTO", -200), ("INSS", -100),
    ("FGTS", -100), ("FERIAS A PAGAR", -50),
    ("RESCISOES A PAGAR", -50), ("FORNECEDORES", -400),
]

# ── Débito para Rescisão (Art.480): Rescisões a Pagar (560) ──────────────
MULTA480_DEBITO_POS = [
    ("RESCISOES A PAGAR", 500),
    ("SALARIOS E ORDENADOS A PAGAR", 300),
    ("OBRIGACOES COM O PESSOAL", 200),
]
MULTA480_DEBITO_NEG = [
    ("DESPESA", -300), ("CUSTO", -300), ("INSS", -200),
    ("FGTS", -200), ("A RECOLHER", -300), ("PROVISAO", -200),
    ("FORNECEDORES", -400),
]

# ── Crédito para Rescisão (Art.480): Indenizações/Aviso Prévio (DRE) ─────
MULTA480_CREDITO_POS = [
    ("INDENIZACOES E AVISO PREVIO", 500),
    ("INDENIZACOES TRABALHISTAS", 500),
    ("MULTAS RESCISORIAS", 500),
    ("DESPESAS COM RESCISAO", 450),
    ("RESCISOES DE CONTRATO", 450),
    ("INDENIZACAO ADICIONAL", 400),
    ("MULTA ESTABILIDADE", 400),
    ("DESPESAS COM PESSOAL", 200),
]
MULTA480_CREDITO_NEG = [
    ("A PAGAR", -300), ("A RECOLHER", -300),
    ("PROVISAO", -300), ("FORNECEDORES", -400),
    ("INSS", -200), ("FGTS", -200),
    ("FERIAS", -200), ("13 SALARIO", -200),
]

# ── Crédito Plano Saúde / Odonto / Coparticipação (DRE) ──────────────────
DESC_PLANO_SAUDE_CRED_POS = [
    ("ASSISTENCIA MEDICA E SOCIAL", 500),
    ("ASSISTENCIA MEDICA", 450),
    ("PLANO DE SAUDE", 400),
    ("PLANO SAUDE", 400),
    ("DESPESAS COM PESSOAL", 100),
]
DESC_PLANO_SAUDE_CRED_NEG = [
    ("A PAGAR", -300), ("A RECOLHER", -300),
    ("A COMPENSAR", -400), ("PROVISAO", -200),
    ("SALARIOS", -150), ("INSS", -100), ("FGTS", -100),
    ("CUSTOS DOS PRODUTOS VENDIDOS", -500),
]

# ── Crédito Vale Transporte (DRE) ─────────────────────────────────────────
DESC_VT_CRED_POS = [
    ("VALE TRANSPORTE", 500),
    ("DESPESAS COM PESSOAL", 100),
]
DESC_VT_CRED_NEG = [
    ("A PAGAR", -300), ("A RECOLHER", -300),
    ("A COMPENSAR", -400), ("PROVISAO", -200),
    ("SALARIOS", -150), ("ALIMENTACAO", -200),
    ("VALE REFEICAO", -200), ("CUSTOS DOS PRODUTOS VENDIDOS", -500),
]

# ── Crédito Vale Refeição / Alimentação (DRE) ─────────────────────────────
DESC_VR_CRED_POS = [
    ("VALE REFEICAO", 500),
    ("DESPESAS COM ALIMENTACAO", 450),
    ("ALIMENTACAO/ CESTA BASICA", 450),
    ("ALIMENTACAO", 400),
    ("DESPESAS COM PESSOAL", 100),
]
DESC_VR_CRED_NEG = [
    ("A PAGAR", -300), ("A RECOLHER", -300),
    ("A COMPENSAR", -400), ("PROVISAO", -200),
    ("SALARIOS", -150), ("VALE TRANSPORTE", -100),
    ("CUSTOS DOS PRODUTOS VENDIDOS", -500),
]

# ── Crédito Adiantamento Salarial: Adiantamento de Salário (Ativo 25) ─────
DESC_ADIANT_SAL_CRED_POS = [
    ("ADIANTAMENTO DE SALARIO", 500),
    ("ADIANTAMENTO DE 13", 450),
    ("ADIANTAMENTOS DE FERIAS", 400),
    ("ADIANTAMENTO A EMPREGADOS", 400),
    ("ADIANTAMENTOS", 200),
]
DESC_ADIANT_SAL_CRED_NEG = [
    ("DESPESA", -200), ("CUSTO", -200),
    ("A RECOLHER", -300), ("INSS", -200), ("FGTS", -200),
    ("PROVISAO", -200), ("SALARIOS A PAGAR", -100),
    ("CUSTOS DOS PRODUTOS VENDIDOS", -500),
]

# ── Débito Adiantamento Salarial (Provento): Adiantamento de Salário (Ativo)
ADIANT_SAL_DEBITO_POS = [
    ("ADIANTAMENTO DE SALARIO", 500),
    ("ADIANTAMENTO A EMPREGADOS", 400),
    ("ADIANTAMENTOS", 200),
]
ADIANT_SAL_DEBITO_NEG = [
    ("DESPESA", -200), ("CUSTO", -200),
    ("A RECOLHER", -300), ("INSS", -200), ("FGTS", -200),
    ("PROVISAO", -200), ("SALARIOS A PAGAR", -100),
    ("FORNECEDORES", -400),
]

# ── Crédito Estouro Mês Anterior (DRE Salários) ───────────────────────────
DESC_ESTOURO_CRED_POS = [
    ("SALARIOS E ORDENADOS", 400),
    ("PRO-LABORE", 300),
    ("DESPESAS COM PESSOAL", 250),
    ("MAO-DE-OBRA DIRETA", 300),
]
DESC_ESTOURO_CRED_NEG = [
    ("A PAGAR", -300), ("A RECOLHER", -300),
    ("A COMPENSAR", -400), ("PROVISAO", -300),
    ("OBRIGACOES", -200), ("INSS", -100), ("FGTS", -100),
    ("HORAS EXTRAS", -100), ("ALIMENTACAO", -300),
    ("VALE REFEICAO", -300), ("CUSTOS DOS PRODUTOS VENDIDOS", -500),
]

# ── Adiantamento de Férias (Desconto): D=Férias a Pagar / C=Adiant.Férias ─
ADIANT_FER_DEBITO_POS = [
    ("FERIAS A PAGAR", 500),
    ("PROVISOES PARA FERIAS", 400),
    ("OBRIGACOES COM O PESSOAL", 250),
]
ADIANT_FER_DEBITO_NEG = [
    ("DESPESA", -300), ("CUSTO", -300),
    ("INSS", -200), ("FGTS", -200),
    ("A RECOLHER", -300), ("PROVISAO", -100),
    ("SALARIOS E ORDENADOS A PAGAR", -100),
    ("RESCISOES A PAGAR", -100),
]
ADIANT_FER_CRED_POS = [
    ("ADIANTAMENTOS DE FERIAS", 500),
    ("ADIANTAMENTO A EMPREGADOS", 400),
    ("ADIANTAMENTOS", 200),
]
ADIANT_FER_CRED_NEG = [
    ("DESPESA", -200), ("CUSTO", -200),
    ("A RECOLHER", -300), ("INSS", -200), ("FGTS", -200),
    ("PROVISAO", -200), ("SALARIOS A PAGAR", -100),
    ("CUSTOS DOS PRODUTOS VENDIDOS", -500),
]

# ── INSS Desconto: D=Salários a Pagar / C=INSS a Recolher ────────────────
DESC_INSS_CRED_POS = [
    ("INSS A RECOLHER", 500),
    ("OBRIGACOES SOCIAIS", 400),
    ("OBRIGACOES TRABALHISTAS E PREVIDENCIARIA", 350),
]
DESC_INSS_CRED_NEG = [
    ("DESPESA", -300), ("CUSTO", -300),
    ("FERIAS", -100), ("13 SALARIO", -100),
    ("PROVISAO", -200), ("FORNECEDORES", -400),
]

# ── IRRF Desconto: D=Salários a Pagar / C=IRRF a Recolher ────────────────
DESC_IRRF_CRED_POS = [
    ("IMPOSTO DE RENDA A RECOLHER", 500),
    ("IRRF A RECOLHER", 500),
    ("IRRF S/ FOLHA", 450),
]
DESC_IRRF_CRED_NEG = [
    ("DESPESA", -300), ("CUSTO", -300),
    ("PROVISAO", -200), ("FORNECEDORES", -400),
    ("INSS", -200), ("FGTS", -200),
]

# ── Sindicato: D=Salários a Pagar / C=Contribuições Sindicais ────────────
DESC_SINDICATO_CRED_POS = [
    ("CONTRIBUICOES SINDICAIS", 500),
    ("CONTRIBUICAO SINDICAL A RECOLHER", 450),
    ("OBRIGACOES SOCIAIS", 300),
]
DESC_SINDICATO_CRED_NEG = [
    ("DESPESA", -300), ("CUSTO", -300),
    ("PROVISAO", -200), ("FORNECEDORES", -400),
]

# ── Pensão Alimentícia: D=Salários a Pagar / C=Pensão a Pagar ────────────
DESC_PENSAO_CRED_POS = [
    ("PENSAO ALIMENTICIA A PAGAR", 500),
    ("OBRIGACOES COM O PESSOAL", 300),
]
DESC_PENSAO_CRED_NEG = [
    ("DESPESA", -300), ("CUSTO", -300),
    ("INSS", -200), ("FGTS", -200),
    ("PROVISAO", -200), ("FORNECEDORES", -400),
]

# ── Consignado: D=Salários a Pagar / C=Empréstimos (Ativo) ───────────────
DESC_CONSIGNADO_CRED_POS = [
    ("EMPRESTIMOS", 500),
    ("ADIANTAMENTOS", 200),
]
DESC_CONSIGNADO_CRED_NEG = [
    ("DESPESA", -200), ("CUSTO", -200),
    ("A RECOLHER", -300), ("INSS", -200), ("FGTS", -200),
    ("PROVISAO", -200), ("SALARIOS A PAGAR", -100),
    ("FORNECEDORES", -400),
]

# ── FGTS Informativa: D=FGTS (DRE) / C=FGTS a Recolher ──────────────────
FGTS_DEBITO_POS = [
    ("FGTS", 500), ("ENCARGOS SOCIAIS", 400),
    ("DESPESAS COM PESSOAL", 300),
]
FGTS_DEBITO_NEG = [
    ("A PAGAR", -300), ("A RECOLHER", -300),
    ("PROVISAO", -200), ("FORNECEDORES", -400),
    ("SALARIOS E ORDENADOS A PAGAR", -100),
]
FGTS_CREDITO_POS = [
    ("FGTS A RECOLHER", 500),
    ("FGTS SOBRE PROVISOES PARA FERIAS", 400),
    ("OBRIGACOES SOCIAIS", 300),
]
FGTS_CREDITO_NEG = [
    ("DESPESA", -300), ("CUSTO", -300),
    ("PROVISAO", -100), ("FORNECEDORES", -400),
]

# ── Multa Estabilidade Art.479 (Provento): D=Indenizações (DRE) / C=Rescisões a Pagar
MULTA479_DEBITO_POS = [
    ("INDENIZACOES E AVISO PREVIO", 500),
    ("INDENIZACOES TRABALHISTAS", 500),
    ("MULTAS RESCISORIAS", 500),
    ("DESPESAS COM RESCISAO", 450),
    ("RESCISOES DE CONTRATO", 450),
    ("INDENIZACAO ADICIONAL", 400),
    ("MULTA ESTABILIDADE", 400),
    ("DESPESAS COM PESSOAL", 200),
]
MULTA479_DEBITO_NEG = [
    ("A PAGAR", -300), ("A RECOLHER", -300),
    ("PROVISAO", -300), ("FORNECEDORES", -400),
    ("INSS", -200), ("FGTS", -200),
    ("FERIAS", -200), ("13 SALARIO", -200),
]
MULTA479_CREDITO_POS = [
    ("RESCISOES A PAGAR", 500),
    ("SALARIOS E ORDENADOS A PAGAR", 400),
    ("OBRIGACOES COM O PESSOAL", 300),
]
MULTA479_CREDITO_NEG = [
    ("DESPESA", -300), ("CUSTO", -300),
    ("INSS", -200), ("FGTS", -200),
    ("A RECOLHER", -300), ("PROVISAO", -200),
    ("FORNECEDORES", -400),
]


# ══════════════════════════════════════════════════════════════════════════
# FUNÇÕES DE DETECÇÃO DE RUBRICAS ESPECÍFICAS
# ══════════════════════════════════════════════════════════════════════════
def _e_plano_saude_odonto(nome: str) -> bool:
    termos = [
        "PLANO SAUDE", "PLANO DE SAUDE", "ODONTOLOGICO", "ODONTO",
        "COPARTICIPACAO", "COPART", "PLANO ODONT", "DESCONTO PLANO",
        "DESC COPARTICIPACAO", "DESC PLANO", "8111", "DESCONTO PLANO DE SAUDE",
        "DESCONTO PLANO ODONTOLOGICOS",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_vale_transporte(nome: str) -> bool:
    termos = [
        "VALE TRANSPORTE", "VALE TRANSP", "VT 6",
        "DESC VT", "DESCONTO VT", "DESCONTO VALE TRANSP",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_vale_refeicao(nome: str) -> bool:
    termos = [
        "DESC VALE REFEICAO", "DESCONTO VALE REFEICAO",
        "DESC. VALE REFEICAO", "DESCONTO VR", "DESC VALE REFEI",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_adiantamento_salarial_desc(nome: str) -> bool:
    """Desconto de adiantamento salarial — crédito vai para Ativo (conta 25)."""
    termos = [
        "DESC.ADIANT.SALARIAL", "DESC ADIANT SALARIAL",
        "DESCONTO ADIANTAMENTO SALARIAL", "DESC.ADIANT",
        "DESC ADIANT",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_adiantamento_salarial_provento(nome: str) -> bool:
    """Provento de adiantamento salarial — D=Adiantamento de Salário (Ativo) / C=Salários a Pagar."""
    termos = [
        "ADIANTAMENTO SALARIAL",
    ]
    # Não pode ser desconto
    nome_n = _norm(nome)
    if any(_norm(t) in nome_n for t in ["DESC", "DESCONTO"]):
        return False
    return any(_norm(t) in nome_n for t in termos)


def _e_adiantamento_ferias_desc(nome: str) -> bool:
    """Desconto de adiantamento de férias."""
    termos = [
        "ADIANTAMENTO DE FERIAS", "ADIANTAMENTO FERIAS",
        "ADIANT DE FERIAS", "ADIANT FERIAS",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_estouro(nome: str) -> bool:
    termos = [
        "ESTOURO MES ANTERIOR", "ESTOURO MES ANT",
        "ESTOURO SEMANA ANTERIOR", "TROCO MES ANTERIOR",
        "TROCO SEMANA ANTERIOR", "TROCO MES ANT", "ESTOURO ANTERIOR",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_inss_desconto(nome: str) -> bool:
    termos = [
        "I.N.S.S.", "INSS FERIAS", "INSS SOBRE RESCISAO",
        "INSS 13 SAL", "INSS DIFERENCA FERIAS",
    ]
    nome_n = _norm(nome)
    # Não pode ser FGTS
    if "FGTS" in nome_n:
        return False
    return any(_norm(t) in nome_n for t in termos)


def _e_irrf(nome: str) -> bool:
    termos = [
        "IMPOSTO DE RENDA", "IRRF FERIAS", "IRRF",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_fgts_informativa(nome: str) -> bool:
    termos = [
        "F.G.T.S DO MES", "F.G.T.S  DO MES", "FGTS DO MES",
        "FGTS FERIAS", "FGTS 13O SALARIO RESCISAO",
        "F.G.T.S DE RESCISAO", "FGTS 13O SALARIO RESCISAO",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_sindicato(nome: str) -> bool:
    termos = [
        "CONTRIBUICAO SINDICAL", "MENSALIDADE SINDICAL",
        "SINDICATO",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_pensao(nome: str) -> bool:
    termos = ["PENSAO ALIMENTICIA"]
    return any(_norm(t) in nome for t in termos)


def _e_consignado(nome: str) -> bool:
    termos = [
        "EMPRESTIMO CONSIGNADO", "EMP. CRED. TRAB",
        "EMP CRED TRAB", "DESC. EMPRESTIMO CONSIGNADO",
    ]
    return any(_norm(t) in nome for t in termos)


def _e_multa_479(nome: str) -> bool:
    """Multa Estabilidade Art.479 (Provento) — empresa paga ao funcionário."""
    termos = [
        "MULTA ESTABILIDADE ART. 479", "MULTA ESTABILIDADE ART.479",
        "MULTA ESTABILIDADE ART 479", "MULTA ATRASO PAGTO",
        "MULTA ATRASO PAGAMENTO", "8115", "8124", "8527",
    ]
    nome_n = _norm(nome)
    # Não pode ser desconto
    if any(_norm(t) in nome_n for t in ["DESC", "DESCONTO", "480"]):
        return False
    # Detecta genérico "MULTA ESTABILIDADE" sem Art.480
    if _norm("MULTA ESTABILIDADE") in nome_n and _norm("480") not in nome_n:
        return True
    return any(_norm(t) in nome_n for t in termos)


def _e_multa_480(nome: str) -> bool:
    """Multa Estabilidade Art.480 (Desconto) — funcionário paga à empresa."""
    termos = [
        "MULTA ESTABILIDADE ART. 480", "MULTA ESTABILIDADE ART.480",
        "MULTA ESTABILIDADE ART 480", "ART 480", "ART. 480",
        "DESC. MULTA ESTABILIDADE", "DESC MULTA ESTABILIDADE",
        "8116", "8528",
    ]
    return any(_norm(t) in nome for t in termos)


# ══════════════════════════════════════════════════════════════════════════
# CORE: GERAR DE/PARA EVENTO vs CONTA CONTÁBIL
# ══════════════════════════════════════════════════════════════════════════
def gerar_depara_evento_conta(
    evento_cod:  str,
    evento_nome: str,
    evento_tipo: str,
    grupo:       str,
    df_contas:   pd.DataFrame,
    tipo_folha:  str = "1",
) -> dict:
    """
    Retorna as contas de débito e crédito para um evento.
    Aplica regras específicas por rubrica antes do fallback por grupo.
    """
    vazio = {"conta_debito": "", "conta_credito": "", "desc_debito": "", "desc_credito": ""}

    if df_contas is None or df_contas.empty:
        return vazio

    nome = _norm(evento_nome)

    # ── Encargo Patronal (Tipo Folha = Empresa) ───────────────────────────
    if tipo_folha == "2":
        df_d, df_c = filtrar_contas_por_grupo(df_contas, "Encargo Patronal", True)

        def _melhor(df_g):
            if df_g.empty:
                return "", ""
            for _, r in df_g.iterrows():
                n = r["nome_conta"]
                if "INSS A RECOLHER" in n or "OBRIGACOES SOCIAIS" in n:
                    return r["reduzido"], r["nome_original"]
            r = df_g.iloc[0]
            return r["reduzido"], r["nome_original"]

        cd, dd = _melhor(df_d)
        cc, dc = _melhor(df_c)
        return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

    # ══════════════════════════════════════════════════════════════════════
    # REGRAS ESPECÍFICAS POR TIPO DE RUBRICA
    # ══════════════════════════════════════════════════════════════════════

    # ── FGTS Informativa ─────────────────────────────────────────────────
    if evento_tipo == "Informativa" and _e_fgts_informativa(nome):
        cd, dd = _melhor_conta_por_score(df_contas, FGTS_DEBITO_POS, FGTS_DEBITO_NEG, score_min=100)
        cc, dc = _melhor_conta_por_score(df_contas, FGTS_CREDITO_POS, FGTS_CREDITO_NEG, score_min=100)
        if cd and cc:
            return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

    # ── PROVENTOS ─────────────────────────────────────────────────────────
    if evento_tipo == "Provento":

        # Adiantamento Salarial (Provento): D=Adiantamento de Salário (Ativo) / C=187
        if _e_adiantamento_salarial_provento(nome):
            cd, dd = _melhor_conta_por_score(df_contas, ADIANT_SAL_DEBITO_POS, ADIANT_SAL_DEBITO_NEG, score_min=50)
            cc, dc = _melhor_conta_por_score(df_contas, DESCONTO_DEBITO_SAL_PAGAR_POS, DESCONTO_DEBITO_SAL_PAGAR_NEG, score_min=100)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # Multa Estabilidade Art.479 (Provento): D=Indenizações (DRE) / C=Rescisões a Pagar
        if _e_multa_479(nome):
            cd, dd = _melhor_conta_por_score(df_contas, MULTA479_DEBITO_POS, MULTA479_DEBITO_NEG, score_min=100)
            cc, dc = _melhor_conta_por_score(df_contas, MULTA479_CREDITO_POS, MULTA479_CREDITO_NEG, score_min=100)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

    # ── DESCONTOS ─────────────────────────────────────────────────────────
    if evento_tipo in ("Desconto", "Inf. Dedutora"):

        # Débito padrão para descontos = Salários e Ordenados a Pagar (187)
        cd, dd = _melhor_conta_por_score(
            df_contas, DESCONTO_DEBITO_SAL_PAGAR_POS, DESCONTO_DEBITO_SAL_PAGAR_NEG, score_min=100
        )

        # ── Adiantamento de Férias: D=Férias a Pagar / C=Adiantamentos de Férias
        if _e_adiantamento_ferias_desc(nome):
            cd, dd = _melhor_conta_por_score(df_contas, ADIANT_FER_DEBITO_POS, ADIANT_FER_DEBITO_NEG, score_min=100)
            cc, dc = _melhor_conta_por_score(df_contas, ADIANT_FER_CRED_POS, ADIANT_FER_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Multa Estabilidade Art.480: D=Rescisões a Pagar / C=Indenizações (DRE)
        elif _e_multa_480(nome):
            cd, dd = _melhor_conta_por_score(df_contas, MULTA480_DEBITO_POS, MULTA480_DEBITO_NEG, score_min=100)
            cc, dc = _melhor_conta_por_score(df_contas, MULTA480_CREDITO_POS, MULTA480_CREDITO_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── INSS: D=187 / C=INSS a Recolher
        elif _e_inss_desconto(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_INSS_CRED_POS, DESC_INSS_CRED_NEG, score_min=100)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── IRRF: D=187 / C=IRRF a Recolher
        elif _e_irrf(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_IRRF_CRED_POS, DESC_IRRF_CRED_NEG, score_min=100)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Sindicato: D=187 / C=Contribuições Sindicais
        elif _e_sindicato(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_SINDICATO_CRED_POS, DESC_SINDICATO_CRED_NEG, score_min=100)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Pensão Alimentícia: D=187 / C=Pensão a Pagar
        elif _e_pensao(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_PENSAO_CRED_POS, DESC_PENSAO_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Consignado: D=187 / C=Empréstimos (Ativo)
        elif _e_consignado(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_CONSIGNADO_CRED_POS, DESC_CONSIGNADO_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Plano Saúde / Odonto / Coparticipação: D=187 / C=Assist.Médica (DRE)
        elif _e_plano_saude_odonto(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_PLANO_SAUDE_CRED_POS, DESC_PLANO_SAUDE_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Vale Transporte: D=187 / C=Vale Transporte (DRE)
        elif _e_vale_transporte(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_VT_CRED_POS, DESC_VT_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Vale Refeição: D=187 / C=Alimentação/VR (DRE)
        elif _e_vale_refeicao(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_VR_CRED_POS, DESC_VR_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Adiantamento Salarial (Desconto): D=187 / C=Adiantamento de Salário (Ativo)
        elif _e_adiantamento_salarial_desc(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_ADIANT_SAL_CRED_POS, DESC_ADIANT_SAL_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # ── Estouro Mês Anterior: D=187 / C=Salários e Ordenados (DRE)
        elif _e_estouro(nome):
            cc, dc = _melhor_conta_por_score(df_contas, DESC_ESTOURO_CRED_POS, DESC_ESTOURO_CRED_NEG, score_min=50)
            if cd and cc:
                return {"conta_debito": cd, "conta_credito": cc, "desc_debito": dd, "desc_credito": dc}

        # Retorna com débito encontrado e crédito vazio (sem regra específica)
        return {"conta_debito": cd, "conta_credito": "", "desc_debito": dd, "desc_credito": ""}

    # ══════════════════════════════════════════════════════════════════════
    # FALLBACK: BUSCA POR GRUPO (Proventos e Informativas sem regra específica)
    # ══════════════════════════════════════════════════════════════════════
    aplicar_filtro = (grupo != "Outro")
    df_d, df_c = filtrar_contas_por_grupo(df_contas, grupo, aplicar_filtro_folha=aplicar_filtro)

    evento_norm   = nome
    palavras_ev   = [p for p in evento_norm.split() if len(p) > 3]

    def _melhor_fallback(df_grupo: pd.DataFrame) -> tuple[str, str]:
        if df_grupo.empty:
            return "", ""
        melhor_score = -1
        melhor_idx   = 0
        for idx, row in df_grupo.iterrows():
            nome_c = row["nome_conta"]
            score  = sum(1 for p in palavras_ev if p in nome_c)
            if "score_folha" in row:
                score += row["score_folha"] * 0.1
            if score > melhor_score:
                melhor_score = score
                melhor_idx   = df_grupo.index.get_loc(idx)
        row_sel = df_grupo.iloc[melhor_idx]
        return row_sel["reduzido"], row_sel["nome_original"]

    cod_deb,  desc_deb  = _melhor_fallback(df_d)
    cod_cred, desc_cred = _melhor_fallback(df_c)

    return {
        "conta_debito":  cod_deb,
        "conta_credito": cod_cred,
        "desc_debito":   desc_deb,
        "desc_credito":  desc_cred,
    }


# ══════════════════════════════════════════════════════════════════════════
# CLASSIFICAÇÃO AUTOMÁTICA
# ══════════════════════════════════════════════════════════════════════════
def classificar_todos_eventos(
    eventos:   list,
    catalog:   dict,
    df_contas: pd.DataFrame | None,
    log:       list,
) -> dict:
    resultado: dict[str, dict] = {}
    chaves_unicas = {(ev["cod"], ev["tipo_folha"]) for ev in eventos}

    for cod, tipo_folha in chaves_unicas:
        info  = catalog.get(cod, {})
        nome  = info.get("descricao", cod)
        tipo  = info.get("tipo", "Provento")

        classif   = classificar_rubrica_local(nome, tipo, tipo_folha)
        grupo     = classif["grupo"]
        confianca = classif["confianca"]

        conta_debito = conta_credito = ""
        desc_debito  = desc_credito  = ""

        if df_contas is not None and not df_contas.empty:
            depara = gerar_depara_evento_conta(cod, nome, tipo, grupo, df_contas, tipo_folha)
            conta_debito  = depara["conta_debito"]
            conta_credito = depara["conta_credito"]
            desc_debito   = depara["desc_debito"]
            desc_credito  = depara["desc_credito"]

        chave_resultado = f"{cod}_{tipo_folha}"
        resultado[chave_resultado] = {
            "grupo":         grupo,
            "confianca":     confianca,
            "conta_debito":  conta_debito,
            "conta_credito": conta_credito,
            "desc_debito":   desc_debito,
            "desc_credito":  desc_credito,
            "tipo_folha":    tipo_folha,
        }

    n_alta   = sum(1 for v in resultado.values() if v["confianca"] == "alta")
    n_media  = sum(1 for v in resultado.values() if v["confianca"] == "media")
    n_baixa  = sum(1 for v in resultado.values() if v["confianca"] == "baixa")
    n_patron = sum(1 for v in resultado.values() if v["grupo"] == "Encargo Patronal")
    log.append(
        f"Classificação automática: {len(resultado)} rubricas → "
        f"🟢 {n_alta} alta · 🟡 {n_media} média · 🔴 {n_baixa} baixa · "
        f"🏛️ {n_patron} Encargo Patronal"
    )
    return resultado


def classificar_eventos_por_grupo_cc(
    eventos:   list,
    catalog:   dict,
    grupo_cc:  str,
    df_contas: pd.DataFrame,
    log:       list,
) -> dict:
    resultado: dict[str, dict] = {}

    for ev in eventos:
        cod        = ev["cod"]
        tipo_folha = ev["tipo_folha"]
        info       = catalog.get(cod, {})
        nome       = info.get("descricao", cod)
        tipo       = info.get("tipo", "Provento")

        grupo_efetivo = "Encargo Patronal" if tipo_folha == "2" else grupo_cc

        depara = gerar_depara_evento_conta(cod, nome, tipo, grupo_efetivo, df_contas, tipo_folha)

        chave = f"{cod}_{tipo_folha}"
        resultado[chave] = {
            "grupo":         grupo_efetivo,
            "confianca":     "manual" if tipo_folha != "2" else "alta",
            "conta_debito":  depara["conta_debito"],
            "conta_credito": depara["conta_credito"],
            "desc_debito":   depara["desc_debito"],
            "desc_credito":  depara["desc_credito"],
            "tipo_folha":    tipo_folha,
        }

    log.append(f"CC com grupo '{grupo_cc}': {len(resultado)} evento(s) classificado(s).")
    return resultado


# ══════════════════════════════════════════════════════════════════════════
# PARSE TXT RUBRICAS
# ══════════════════════════════════════════════════════════════════════════
def parse_rubricas_txt(file_bytes: bytes, log: list) -> dict:
    catalog = {}
    TIPO_MAP = {"P": "Provento", "D": "Desconto", "I": "Informativa", "ID": "Inf. Dedutora"}
    try:
        texto = file_bytes.decode("latin-1", errors="replace")
    except Exception as e:
        log.append(f"ERRO ao decodificar rubricas.txt: {e}")
        return catalog

    for raw in texto.splitlines():
        raw = raw.strip()
        if not raw:
            continue
        partes = raw.split("\t")
        if len(partes) < 5:
            continue
        cod       = partes[2].strip()
        descricao = partes[3].strip()
        tipo_raw  = partes[4].strip().upper()
        if not cod:
            continue
        tipo_norm = TIPO_MAP.get(tipo_raw)
        if tipo_norm is None:
            continue
        if cod not in catalog:
            catalog[cod] = {"tipo": tipo_norm, "descricao": descricao}

    log.append(f"rubricas.txt: {len(catalog)} código(s) mapeado(s).")
    return catalog


# ══════════════════════════════════════════════════════════════════════════
# PARSE PDF
# ══════════════════════════════════════════════════════════════════════════
IGNORE_PATTERNS = [
    r"^RELAÇÃO DE RUBRICAS", r"^Página", r"^Emissão",
    r"^Hora:", r"^Empresa:", r"^Código\s+Descrição", r"^\s*$",
]
SECAO_TIPO_FOLHA = {
    "Folha Normal": "1", "Empresa": "2", "Férias": "3",
    "Rescisão": "4", "Provisão de Férias": "5",
    "Provisão de 13º": "6", "Provisão de 13o": "6",
}
SECAO_TIPO_FOLHA_DESC = {
    "1": "Folha Normal", "2": "Empresa", "3": "Férias",
    "4": "Rescisão", "5": "Provisão de Férias", "6": "Provisão de 13º",
}
RE_SECAO = re.compile(
    r"^(Folha Normal|Empresa|Férias|Rescisão|"
    r"Provisão de Férias|Provisão de 13º|Provisão de 13o)$",
    re.IGNORECASE,
)
RE_CC    = re.compile(r"^Centro de Custo:\s*(\d+)\s+(.+)$", re.IGNORECASE)
RE_EVENT = re.compile(r"^\s*(\d+)\s+(.+)$")


def should_ignore(line: str) -> bool:
    return any(re.search(p, line, re.IGNORECASE) for p in IGNORE_PATTERNS)


def parse_nao_configurados_pdf(file_bytes: bytes, log: list) -> list:
    eventos = []
    vistos  = set()
    tipo_folha_atual = "1"
    cc_cod_atual = cc_nome_atual = ""

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for raw in text.splitlines():
                line = raw.strip()
                if not line:
                    continue
                m = RE_SECAO.match(line)
                if m:
                    sec = m.group(1).strip()
                    for k, v in SECAO_TIPO_FOLHA.items():
                        if k.lower() in sec.lower():
                            tipo_folha_atual = v
                            break
                    continue
                m = RE_CC.match(line)
                if m:
                    cc_cod_atual  = m.group(1).strip()
                    cc_nome_atual = m.group(2).strip()
                    continue
                if should_ignore(line):
                    continue
                m = RE_EVENT.match(line)
                if m:
                    cod  = m.group(1).strip()
                    desc = m.group(2).strip()
                    if not cod.isdigit():
                        continue
                    chave = (cod, tipo_folha_atual, cc_cod_atual)
                    if chave not in vistos:
                        vistos.add(chave)
                        eventos.append({
                            "cod":               cod,
                            "descricao_pdf":     desc,
                            "tipo_folha":        tipo_folha_atual,
                            "tipo_folha_desc":   SECAO_TIPO_FOLHA_DESC.get(tipo_folha_atual, tipo_folha_atual),
                            "centro_custo_cod":  cc_cod_atual,
                            "centro_custo_nome": cc_nome_atual,
                        })

    por_folha: dict[str, int] = {}
    for ev in eventos:
        desc = ev["tipo_folha_desc"]
        por_folha[desc] = por_folha.get(desc, 0) + 1
    resumo = " · ".join(f"{k}: {v}" for k, v in por_folha.items())
    log.append(f"PDF: {len(eventos)} evento(s) extraído(s). [{resumo}]")
    return eventos


def get_centros_custo_unicos(eventos: list) -> list[tuple[str, str]]:
    vistos: dict[str, str] = {}
    for ev in eventos:
        cod  = ev["centro_custo_cod"]
        nome = ev["centro_custo_nome"]
        if cod and cod not in vistos:
            vistos[cod] = nome
    return list(vistos.items())


def get_eventos_por_cc(eventos: list, cc_cod: str) -> list:
    return [ev for ev in eventos if ev["centro_custo_cod"] == cc_cod]


# ══════════════════════════════════════════════════════════════════════════
# GERA EXCEL — ETAPA 1
# ══════════════════════════════════════════════════════════════════════════
def gerar_excel_configuracao(
    eventos:       list,
    catalog:       dict,
    cod_empresa:   str,
    log:           list,
    usa_separador: bool = False,
    config_cc:     dict | None = None,
    df_contas:     pd.DataFrame | None = None,
    classif_auto:  dict | None = None,
) -> bytes:
    linhas = []

    for ev in eventos:
        cod        = ev["cod"]
        tipo_folha = ev["tipo_folha"]
        info       = catalog.get(cod, {})
        tipo       = info.get("tipo", "⚠️ Não encontrado")
        desc_rubr  = info.get("descricao", ev["descricao_pdf"])
        cc_cod     = ev["centro_custo_cod"]

        conta_deb = conta_cred = historico = grupo = ""
        desc_deb  = desc_cred = ""

        chave_auto = f"{cod}_{tipo_folha}"

        if usa_separador and config_cc and cc_cod in config_cc:
            cfg       = config_cc[cc_cod]
            historico = cfg.get("historico", "")
            grupo     = "Encargo Patronal" if tipo_folha == "2" else cfg.get("grupo", "")

            if df_contas is not None and not df_contas.empty and grupo:
                depara = gerar_depara_evento_conta(
                    cod, desc_rubr, tipo, grupo, df_contas, tipo_folha
                )
                conta_deb  = depara["conta_debito"]
                conta_cred = depara["conta_credito"]
                desc_deb   = depara["desc_debito"]
                desc_cred  = depara["desc_credito"]

        elif classif_auto and chave_auto in classif_auto:
            auto       = classif_auto[chave_auto]
            grupo      = auto.get("grupo", "")
            conta_deb  = auto.get("conta_debito", "")
            conta_cred = auto.get("conta_credito", "")
            desc_deb   = auto.get("desc_debito", "")
            desc_cred  = auto.get("desc_credito", "")

        linhas.append({
            "Cód. Empresa":              cod_empresa,
            "Cód. Evento":               cod,
            "Descrição (PDF)":           ev["descricao_pdf"],
            "Descrição (Rubricas)":      desc_rubr,
            "Tipo Rubrica":              tipo,
            "Tipo Folha (Nº)":           tipo_folha,
            "Tipo Folha":                ev["tipo_folha_desc"],
            "Cód. Centro de Custo":      cc_cod,
            "Centro de Custo":           ev["centro_custo_nome"],
            "Grupo de Despesa":          grupo,
            "Usa Separador":             "Sim" if usa_separador else "Não",
            "Conta Débito":              conta_deb,
            "Descrição Conta Débito":    desc_deb,
            "Conta Crédito":             conta_cred,
            "Descrição Conta Crédito":   desc_cred,
            "Cód. Histórico":            "",
            "Histórico":                 historico,
            "Observação":                "",
        })

    df = pd.DataFrame(linhas)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Configuração", index=False)
        _formatar_planilha_config(writer.sheets["Configuração"], df)

        if df_contas is not None and not df_contas.empty:
            df_exp = df_contas[["reduzido","classificacao","nome_original","tipo","score_folha"]].copy()
            df_exp.columns = ["Código Reduzido","Classificação","Nome da Conta","Tipo (S/A)","Score Folha"]
            df_exp.to_excel(writer, sheet_name="Plano de Contas", index=False)
            _formatar_planilha_saida(writer.sheets["Plano de Contas"])

    output.seek(0)
    log.append(f"Excel gerado: {len(linhas)} linha(s). Separador: {'Sim' if usa_separador else 'Não'}.")
    return output.read()


def _formatar_planilha_config(ws, df: pd.DataFrame):
    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    larguras = {
        "A": 12, "B": 12, "C": 38, "D": 38, "E": 16,
        "F": 14, "G": 20, "H": 18, "I": 22, "J": 22,
        "K": 14, "L": 16, "M": 38, "N": 16, "O": 38,
        "P": 14, "Q": 42, "R": 30,
    }
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    COLS_EDITAVEIS = {12, 14, 16, 17, 18}
    COLS_AUTO      = {13, 15}
    COLS_INFO      = {10, 11}

    TIPO_COR = {
        "Provento":      "D4EDDA",
        "Desconto":      "F8D7DA",
        "Informativa":   "CCE5FF",
        "Inf. Dedutora": "FFF3CD",
    }
    COR_EMPRESA = "E8D5FF"

    for col_idx, cell in enumerate(ws[1], start=1):
        if col_idx in COLS_EDITAVEIS:
            cell.fill = PatternFill("solid", fgColor="FF8000")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        elif col_idx in COLS_AUTO:
            cell.fill = PatternFill("solid", fgColor="28A745")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        elif col_idx in COLS_INFO:
            cell.fill = PatternFill("solid", fgColor="6C757D")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        else:
            cell.fill = PatternFill("solid", fgColor="444444")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda
    ws.row_dimensions[1].height = 32

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        tipo_val       = ws.cell(row=row_idx, column=5).value or ""
        tipo_folha_val = ws.cell(row=row_idx, column=6).value or ""

        if str(tipo_folha_val).strip() == "2":
            cor_linha = COR_EMPRESA
        else:
            cor_linha = TIPO_COR.get(tipo_val, "E2E3E5")

        for col_idx, cell in enumerate(row, start=1):
            cell.border = borda
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in COLS_EDITAVEIS:
                cell.fill = PatternFill("solid", fgColor="FFF8F0")
                cell.font = Font(size=10)
            elif col_idx in COLS_AUTO:
                cell.fill = PatternFill("solid", fgColor="F0FFF4")
                cell.font = Font(size=10, italic=True)
            else:
                cell.fill = PatternFill("solid", fgColor=cor_linha)
                cell.font = Font(size=10)
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ══════════════════════════════════════════════════════════════════════════
# GERA ARQUIVOS FINAIS — ETAPA 2
# ══════════════════════════════════════════════════════════════════════════
def ler_excel_preenchido(file_bytes: bytes, log: list) -> pd.DataFrame | None:
    try:
        xls = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as e:
        log.append(f"ERRO ao abrir Excel preenchido: {e}")
        return None

    sheet = None
    for c in ["Configuração", "configuracao", "Plan1", "Sheet1"]:
        if c in xls.sheet_names:
            sheet = c
            break
    if not sheet:
        sheet = xls.sheet_names[0]

    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, dtype=str)
    except Exception as e:
        log.append(f"ERRO ao ler aba '{sheet}': {e}")
        return None

    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all")
    log.append(f"Excel preenchido: {len(df)} linha(s) na aba '{sheet}'.")
    return df


def _limpa(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none", "") else s


def gerar_arquivos_finais(df: pd.DataFrame, cod_empresa_padrao: str, log: list) -> tuple[bytes, bytes]:
    col_map: dict[str, str] = {}
    for col in df.columns:
        cl = col.lower()
        if   "cód. empresa"              in cl or "cod. empresa"    in cl: col_map["empresa"]          = col
        elif "cód. evento"               in cl or "cod. evento"     in cl: col_map["seq"]              = col
        elif "tipo folha (nº)"           in cl or "tipo folha (n"   in cl: col_map["tipo"]             = col
        elif "descrição (rubricas)"      in cl:                             col_map["desc"]             = col
        elif "descrição (pdf)"           in cl and "desc" not in col_map:  col_map["desc"]             = col
        elif "cód. centro de custo"      in cl:                             col_map["cc"]               = col
        elif "conta débito"              in cl or "conta debito"    in cl: col_map["debito"]           = col
        elif "descrição conta débito"    in cl or "descricao conta debito" in cl: col_map["desc_deb"]  = col
        elif "conta crédito"             in cl or "conta credito"   in cl: col_map["credito"]          = col
        elif "descrição conta crédito"   in cl or "descricao conta credito" in cl: col_map["desc_cred"]= col
        elif "cód. histórico"            in cl or "cod. historico"  in cl: col_map["historico"]        = col
        elif "histórico"                 in cl and "cód" not in cl and "cod" not in cl:
            col_map["historico_texto"] = col
        elif "observação"                in cl:                             col_map["observacao"]       = col
        elif "usa separador"             in cl:                             col_map["usa_separador"]    = col

    TIPO_COL = (
        "Tipo da Integração (1 - Folha mensal; 2 - Empresa; "
        "3 - Férias; 4 - Rescisao; 5 - Prov. Férias; 6 - Prov. 13)"
    )

    linhas_evento, linhas_integra, linhas_integra_xls = [], [], []
    sem_conta = com_conta = 0

    for _, row in df.iterrows():
        empresa     = _limpa(row.get(col_map.get("empresa",      ""), "")) or cod_empresa_padrao
        seq         = _limpa(row.get(col_map.get("seq",          ""), ""))
        tipo        = _limpa(row.get(col_map.get("tipo",         ""), ""))
        desc        = _limpa(row.get(col_map.get("desc",         ""), ""))
        cc          = _limpa(row.get(col_map.get("cc",           ""), ""))
        debito      = _limpa(row.get(col_map.get("debito",       ""), ""))
        credito     = _limpa(row.get(col_map.get("credito",      ""), ""))
        historico   = _limpa(row.get(col_map.get("historico",    ""), ""))
        complemento = _limpa(row.get(col_map.get("historico_texto", ""), ""))
        usa_sep     = _limpa(row.get(col_map.get("usa_separador",""), ""))

        if not seq:
            continue

        sep_val = "1" if usa_sep.lower() == "sim" else "0"
        if debito or credito:
            com_conta += 1
        else:
            sem_conta += 1

        linhas_evento.append({
            "Código da Empresa":               empresa,
            "Centro de custo":                 cc,
            "Código Sequencial da Integração": seq,
            TIPO_COL:                          tipo,
            "Descrição":                       desc,
            "Código da Conta Débito":          debito,
            "Código da Conta Crédito":         credito,
            "Código do Histórico":             historico,
            "Complemento":                     complemento,
        })
        linhas_integra.append({
            "Código da Empresa":               empresa,
            "Separador":                       sep_val,
            "Código Sequencial da Integração": seq,
            TIPO_COL:                          tipo,
            "Código da Rúbrica Selecionada":   seq,
        })
        linhas_integra_xls.append({
            "Código da Empresa":               empresa,
            "Centro de Custo":                 cc,
            "Código Sequencial da Integração": seq,
            TIPO_COL:                          tipo,
            "Descrição":                       desc,
            "Código da Conta Crédito":         credito,
            "Código da Conta Débito":          debito,
            "Código do Histórico":             historico,
        })

    log.append(f"Arquivos → Com conta: {com_conta} | Sem conta: {sem_conta}")

    buf_evento = BytesIO()
    with pd.ExcelWriter(buf_evento, engine="openpyxl") as writer:
        pd.DataFrame(linhas_integra).to_excel(writer, sheet_name="integra", index=False)
        pd.DataFrame(linhas_evento).to_excel(writer,  sheet_name="evento",  index=False)
        for sn in ["integra", "evento"]:
            _formatar_planilha_saida(writer.sheets[sn])
    buf_evento.seek(0)

    buf_integra = BytesIO()
    with pd.ExcelWriter(buf_integra, engine="openpyxl") as writer:
        pd.DataFrame(linhas_integra_xls).to_excel(writer, sheet_name="Plan1", index=False)
        _formatar_planilha_saida(writer.sheets["Plan1"])
    buf_integra.seek(0)

    return buf_evento.read(), buf_integra.read()


def _formatar_planilha_saida(ws):
    borda = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="444444")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda
    ws.row_dimensions[1].height = 32

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = borda
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=10)

    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════
# INTERFACE STREAMLIT — MAIN
# ══════════════════════════════════════════════════════════════════════════
def main():
    st.set_page_config(
        page_title="Domínio | Integração Contábil",
        page_icon="🟠",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    apply_tr_theme()

    st.markdown(
        f"""
        <div style="background:#444444; padding:24px 28px 18px 28px;
                    border-radius:8px; border-top:6px solid #FF8000;
                    margin-bottom:28px;">
            <h2 style="color:#FF8000; margin:0;">
                📊 Integração Contábil — Domínio Sistemas &nbsp;|&nbsp; {VERSAO}
            </h2>
            <p style="color:#DDDDDD; margin:6px 0 0 0;">
                <b>Etapa 1:</b> PDF + TXT + Plano de Contas → classifica automaticamente → gera Excel.<br>
                <b>Etapa 2:</b> Excel preenchido → gera <b>evento exemplo.xlsx</b> e <b>integra exemplo.xlsx</b>.<br>
                <span style="color:#CC99FF;">🏛️ Tipo Folha "Empresa" = INSS Patronal → grupo <b>Encargo Patronal</b> automático.</span>
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Sidebar ────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### ⚙️ Configurações")
        cod_empresa = st.text_input("Código da empresa", value="45")
        st.markdown("---")
        st.markdown("### 🎨 Legenda de Tipos")
        st.markdown("🟢 Verde → Provento")
        st.markdown("🔴 Vermelho → Desconto")
        st.markdown("🔵 Azul → Informativa")
        st.markdown("🟡 Amarelo → Inf. Dedutora")
        st.markdown("🟣 Lilás → Encargo Patronal")
        st.markdown("🟠 Laranja → Campos editáveis")
        st.markdown("🌿 Verde claro → Preenchimento automático")
        st.markdown("---")
        st.markdown("### ℹ️ Regras De/Para")
        st.markdown(
            "**217 VT 6%:** D=187 / C=Vale Transporte (DRE)\n\n"
            "**222/271/8111 Plano/Odonto/Copart:** D=187 / C=Assist.Médica (DRE)\n\n"
            "**260 Vale Refeição:** D=187 / C=Alimentação (DRE)\n\n"
            "**981 Desc.Adiant.Sal:** D=187 / C=25 Adiant.Salário (Ativo)\n\n"
            "**980 Adiant.Salarial:** D=25 Adiant.Salário (Ativo) / C=187\n\n"
            "**991 Estouro Mês Ant:** D=187 / C=Salários e Ordenados (DRE)\n\n"
            "**937 Adiant.Férias:** D=559 Férias a Pagar / C=606 Adiant.Férias\n\n"
            "**842 Multa 480/CLT:** D=560 Rescisões a Pagar / C=281 Indenizações (DRE)\n\n"
            "**831 Multa 479/CLT:** D=281 Indenizações (DRE) / C=560 Rescisões a Pagar\n\n"
            "**Tipo Folha Empresa:** Encargo Patronal automático."
        )
        st.markdown("---")
        st.markdown(f"**Versão:** {VERSAO}")

    # ── Session state ──────────────────────────────────────────────────────
    _defaults = {
        "log":            [f"Pronto. Versão {VERSAO}"],
        "excel_config":   None,
        "evento_xlsx":    None,
        "integra_xls":    None,
        "df_preview":     None,
        "n_eventos":      0,
        "df_contas":      None,
        "eventos_parsed": None,
        "catalog_parsed": None,
        "config_cc":      {},
        "classif_auto":   {},
        "_contas_fid":    None,
        "_contas_name":   None,
    }
    for k, v in _defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # ══════════════════════════════════════════════════════════════════════
    # ETAPA 1
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("## 📋 Etapa 1 — Gerar Excel para Preenchimento")

    col1, col2, col3 = st.columns(3)
    with col1:
        pdf_file = st.file_uploader("1️⃣ PDF — Rubricas/Itens Não Configurados", type=["pdf"], key="pdf_etapa1")
    with col2:
        txt_file = st.file_uploader("2️⃣ TXT — Rubricas (catálogo de tipos)", type=["txt"], key="txt_etapa1")
    with col3:
        contas_file = st.file_uploader(
            "3️⃣ XLS/XLSX — Plano de Contas (opcional)",
            type=["xls", "xlsx"], key="contas_etapa1",
            help="Exportação Plano de Contas - Completo (Kaph Numeric)",
        )

    # Carrega plano de contas
    if contas_file is not None:
        fid   = getattr(contas_file, "file_id", id(contas_file))
        fname = contas_file.name
        if st.session_state._contas_fid != fid:
            log_tmp: list[str] = []
            raw_bytes = contas_file.read()
            df_c = parse_plano_contas(raw_bytes, fname, log_tmp)
            st.session_state.df_contas    = df_c if not df_c.empty else None
            st.session_state._contas_fid  = fid
            st.session_state._contas_name = fname
            st.session_state.config_cc    = {}
            st.session_state.log.extend(log_tmp)
    else:
        if st.session_state._contas_fid is not None:
            st.session_state.df_contas    = None
            st.session_state._contas_fid  = None
            st.session_state._contas_name = None
            st.session_state.config_cc    = {}

    df_pc = st.session_state.df_contas

    if df_pc is not None and not df_pc.empty:
        n_a = len(df_pc[df_pc["tipo"] == "A"])
        n_s = len(df_pc[df_pc["tipo"] == "S"])
        n_folha = len(df_pc[(df_pc["tipo"] == "A") & (df_pc.get("score_folha", pd.Series(dtype=int)) >= SCORE_MINIMO_FOLHA)]) if "score_folha" in df_pc.columns else 0
        st.success(
            f"✅ Plano de Contas **{st.session_state._contas_name}** carregado: "
            f"**{len(df_pc)}** contas ({n_a} analíticas · {n_s} sintéticas · "
            f"**{n_folha}** de folha)"
        )
        with st.expander("🔍 Ver amostra das contas analíticas de folha", expanded=False):
            if "score_folha" in df_pc.columns:
                df_am = df_pc[(df_pc["tipo"] == "A") & (df_pc["score_folha"] >= SCORE_MINIMO_FOLHA)][
                    ["reduzido", "classificacao", "nome_original", "score_folha"]
                ].sort_values("score_folha", ascending=False).head(30)
                df_am.columns = ["Código Reduzido", "Classificação", "Nome da Conta", "Score Folha"]
            else:
                df_am = df_pc[df_pc["tipo"] == "A"][["reduzido", "classificacao", "nome_original"]].head(30)
                df_am.columns = ["Código Reduzido", "Classificação", "Nome da Conta"]
            st.dataframe(df_am, use_container_width=True)
    elif contas_file is not None:
        st.error("❌ Não foi possível carregar o Plano de Contas.")

    st.markdown("---")

    # ── Configuração de Separador ──────────────────────────────────────────
    st.markdown("### ⚙️ Configuração de Separador")
    usa_separador = st.radio(
        "Os lançamentos usam separador por Centro de Custo?",
        ["Não", "Sim"], index=0, horizontal=True,
    )
    usa_sep_bool = (usa_separador == "Sim")

    # ── Configuração por CC ────────────────────────────────────────────────
    if usa_sep_bool:
        if st.session_state.eventos_parsed:
            ccs = get_centros_custo_unicos(st.session_state.eventos_parsed)
            if ccs:
                st.markdown("#### 🏢 Grupo de Despesa por Centro de Custo")
                st.info(
                    "💡 Selecione o **Grupo de Despesa** de cada Centro de Custo. "
                    "Eventos com **Tipo Folha = Empresa** usam automaticamente o grupo "
                    "**Encargo Patronal**, independente do grupo selecionado."
                )

                nao_classif = [
                    f"CC {cc} — {nm}"
                    for cc, nm in ccs
                    if not st.session_state.config_cc.get(cc, {}).get("grupo")
                ]
                if nao_classif:
                    with st.expander(f"⚠️ {len(nao_classif)} CC(s) sem grupo definido", expanded=True):
                        for item in nao_classif:
                            st.markdown(f"- {item}")
                else:
                    st.success("✅ Todos os Centros de Custo têm grupo definido!")

                st.markdown("---")

                for cc_cod, cc_nome in ccs:
                    cfg_atual = st.session_state.config_cc.get(cc_cod, {})
                    grupo_ok  = bool(cfg_atual.get("grupo"))
                    status    = "✅" if grupo_ok else "⚠️"

                    evs_cc    = get_eventos_por_cc(st.session_state.eventos_parsed, cc_cod)
                    n_evs     = len(evs_cc)
                    n_empresa = sum(1 for ev in evs_cc if ev["tipo_folha"] == "2")

                    titulo_cc = f"{status} CC {cc_cod} — {cc_nome} ({n_evs} evento(s)"
                    if n_empresa > 0:
                        titulo_cc += f" · 🏛️ {n_empresa} Encargo Patronal"
                    titulo_cc += ")"

                    with st.expander(titulo_cc, expanded=not grupo_ok):
                        if n_empresa > 0:
                            st.info(
                                f"🏛️ **{n_empresa} evento(s)** deste CC são do Tipo Folha **Empresa** "
                                f"(INSS Patronal) e serão classificados automaticamente como "
                                f"**Encargo Patronal**."
                            )

                        grupo_idx = (
                            GRUPOS_LISTA.index(cfg_atual.get("grupo", "Despesa Administrativa"))
                            if cfg_atual.get("grupo") in GRUPOS_LISTA else 0
                        )
                        grupo_sel = st.selectbox(
                            "📂 Grupo de Despesa do CC (para Folha Normal/Férias/Rescisão)",
                            options=GRUPOS_LISTA,
                            index=grupo_idx,
                            key=f"grupo_{cc_cod}",
                        )
                        hist_sel = st.text_input(
                            "📋 Histórico padrão",
                            value=cfg_atual.get("historico", ""),
                            key=f"hist_{cc_cod}",
                            placeholder="Ex: 001",
                        )

                        if df_pc is not None and not df_pc.empty and grupo_sel != "Outro":
                            col_prev1, col_prev2 = st.columns(2)
                            with col_prev1:
                                df_d, _ = filtrar_contas_por_grupo(df_pc, grupo_sel)
                                n_deb   = len(df_d)
                                if n_deb > 0:
                                    st.success(f"💸 **{n_deb}** conta(s) de débito no grupo")
                                    with st.expander("Ver contas de débito", expanded=False):
                                        for _, r in df_d.head(10).iterrows():
                                            score_txt = f" [score:{r.get('score_folha',0)}]" if "score_folha" in r else ""
                                            st.markdown(f"- `{r['reduzido']}` — {r['nome_original']}{score_txt}")
                                else:
                                    st.warning("⚠️ Nenhuma conta de débito encontrada")
                            with col_prev2:
                                _, df_c = filtrar_contas_por_grupo(df_pc, grupo_sel)
                                n_cred  = len(df_c)
                                if n_cred > 0:
                                    st.success(f"💰 **{n_cred}** conta(s) de crédito no grupo")
                                    with st.expander("Ver contas de crédito", expanded=False):
                                        for _, r in df_c.head(10).iterrows():
                                            score_txt = f" [score:{r.get('score_folha',0)}]" if "score_folha" in r else ""
                                            st.markdown(f"- `{r['reduzido']}` — {r['nome_original']}{score_txt}")
                                else:
                                    st.warning("⚠️ Nenhuma conta de crédito encontrada")

                        if evs_cc and st.session_state.catalog_parsed:
                            with st.expander(f"📋 Ver {n_evs} evento(s) deste CC", expanded=False):
                                dados_evs = []
                                for ev in evs_cc[:25]:
                                    info_ev   = st.session_state.catalog_parsed.get(ev["cod"], {})
                                    grupo_ev  = "🏛️ Encargo Patronal" if ev["tipo_folha"] == "2" else grupo_sel
                                    dados_evs.append({
                                        "Código":        ev["cod"],
                                        "Descrição":     ev["descricao_pdf"],
                                        "Tipo":          info_ev.get("tipo", "—"),
                                        "Tipo Folha":    ev["tipo_folha_desc"],
                                        "Grupo Efetivo": grupo_ev,
                                    })
                                st.dataframe(pd.DataFrame(dados_evs), use_container_width=True)

                        st.session_state.config_cc[cc_cod] = {
                            "grupo":     grupo_sel,
                            "historico": hist_sel,
                        }
        else:
            st.info("⬆️ Faça upload do PDF e clique em **▶ Gerar Excel** para configurar os CCs.")

    st.markdown("---")

    # ── Botões ────────────────────────────────────────────────────────────
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        gerar_excel = st.button(
            "▶ Gerar Excel de Configuração",
            disabled=(pdf_file is None or txt_file is None),
            use_container_width=True,
            type="primary",
        )
    with col_btn2:
        if st.button("🗑 Limpar tudo", use_container_width=True):
            for k in [
                "log","excel_config","evento_xlsx","integra_xls",
                "df_preview","n_eventos","df_contas","eventos_parsed",
                "catalog_parsed","config_cc","classif_auto",
                "_contas_fid","_contas_name",
            ]:
                if k == "log":
                    st.session_state[k] = ["Campos limpos."]
                elif k == "n_eventos":
                    st.session_state[k] = 0
                elif k in ("config_cc", "classif_auto"):
                    st.session_state[k] = {}
                else:
                    st.session_state[k] = None
            st.rerun()

    # ── Processamento ─────────────────────────────────────────────────────
    if gerar_excel and pdf_file and txt_file:
        log: list[str] = ["[Etapa 1] Iniciando..."]

        with st.spinner("Lendo rubricas.txt..."):
            catalog = parse_rubricas_txt(txt_file.read(), log)

        with st.spinner("Lendo PDF..."):
            eventos = parse_nao_configurados_pdf(pdf_file.read(), log)

        st.session_state.eventos_parsed = eventos
        st.session_state.catalog_parsed = catalog

        with st.spinner("🔍 Classificando rubricas automaticamente..."):
            classif_auto = classificar_todos_eventos(eventos, catalog, df_pc, log)
            st.session_state.classif_auto = classif_auto

        if usa_sep_bool:
            ccs_novos = get_centros_custo_unicos(eventos)
            for cc_cod, _ in ccs_novos:
                if cc_cod not in st.session_state.config_cc:
                    evs_cc = [ev for ev in eventos if ev["centro_custo_cod"] == cc_cod and ev["tipo_folha"] != "2"]
                    grupos_cc = [
                        classif_auto.get(f"{ev['cod']}_{ev['tipo_folha']}", {}).get("grupo", "Despesa Administrativa")
                        for ev in evs_cc
                    ]
                    grupo_dom = max(set(grupos_cc), key=grupos_cc.count) if grupos_cc else "Despesa Administrativa"
                    st.session_state.config_cc[cc_cod] = {"grupo": grupo_dom, "historico": ""}
                    log.append(f"CC {cc_cod}: grupo sugerido → {grupo_dom}")

        if not eventos:
            log.append("AVISO: Nenhum evento encontrado no PDF.")
        else:
            with st.spinner("Gerando Excel..."):
                excel_bytes = gerar_excel_configuracao(
                    eventos, catalog, cod_empresa, log,
                    usa_separador=usa_sep_bool,
                    config_cc=st.session_state.config_cc if usa_sep_bool else None,
                    df_contas=df_pc,
                    classif_auto=classif_auto,
                )
            st.session_state.excel_config = excel_bytes
            st.session_state.n_eventos    = len(eventos)

            # Preview
            linhas_prev = []
            for ev in eventos:
                cod_ev     = ev["cod"]
                tipo_folha = ev["tipo_folha"]
                info_ev    = catalog.get(cod_ev, {})
                cc_cod_ev  = ev["centro_custo_cod"]
                chave_auto = f"{cod_ev}_{tipo_folha}"

                if usa_sep_bool and cc_cod_ev in st.session_state.config_cc:
                    cfg_cc = st.session_state.config_cc[cc_cod_ev]
                    grupo  = "Encargo Patronal" if tipo_folha == "2" else cfg_cc.get("grupo", "")
                    if df_pc is not None and not df_pc.empty and grupo:
                        depara = gerar_depara_evento_conta(
                            cod_ev, info_ev.get("descricao", ev["descricao_pdf"]),
                            info_ev.get("tipo", "Provento"), grupo, df_pc, tipo_folha
                        )
                        conta_deb  = depara["conta_debito"]
                        conta_cred = depara["conta_credito"]
                    else:
                        conta_deb = conta_cred = ""
                    confianca = "manual" if tipo_folha != "2" else "alta"
                else:
                    auto       = classif_auto.get(chave_auto, {})
                    grupo      = auto.get("grupo", "—")
                    conta_deb  = auto.get("conta_debito", "")
                    conta_cred = auto.get("conta_credito", "")
                    confianca  = auto.get("confianca", "")

                ok = bool(conta_deb and conta_cred)
                linhas_prev.append({
                    "Código":        cod_ev,
                    "Descrição":     ev["descricao_pdf"],
                    "Tipo":          info_ev.get("tipo", "⚠️"),
                    "Tipo Folha":    ev["tipo_folha_desc"],
                    "Centro Custo":  ev["centro_custo_nome"],
                    "Grupo":         grupo,
                    "Confiança":     confianca,
                    "Conta Débito":  conta_deb,
                    "Conta Crédito": conta_cred,
                    "Classif.":      "✅" if ok else "⚠️",
                })
            st.session_state.df_preview = pd.DataFrame(linhas_prev)

        st.session_state.log = log
        st.rerun()

    # ── Resultado Etapa 1 ──────────────────────────────────────────────────
    if st.session_state.excel_config is not None:
        st.success(f"✅ Excel gerado — {st.session_state.n_eventos} evento(s)")
        st.download_button(
            label="⬇ Baixar Excel de Configuração",
            data=st.session_state.excel_config,
            file_name="configuracao_rubricas_dominio.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

        if st.session_state.df_preview is not None:
            df = st.session_state.df_preview
            total    = len(df)
            p        = len(df[df["Tipo"] == "Provento"])
            d        = len(df[df["Tipo"] == "Desconto"])
            i        = len(df[df["Tipo"] == "Informativa"])
            id_      = len(df[df["Tipo"] == "Inf. Dedutora"])
            nf       = len(df[df["Tipo"].str.startswith("⚠️", na=False)])
            ok       = len(df[df["Classif."] == "✅"]) if "Classif." in df.columns else 0
            nok      = len(df[df["Classif."] == "⚠️"]) if "Classif." in df.columns else 0
            n_patron = len(df[df["Grupo"] == "Encargo Patronal"]) if "Grupo" in df.columns else 0

            cols_m = st.columns(9)
            for col_m, lbl, val in zip(cols_m, [
                "📋 Total","🟢 Proventos","🔴 Descontos","🔵 Informativas",
                "🟡 Inf.Ded.","⚠️ Tipo n/id","✅ Com conta","⚠️ Sem conta","🏛️ Patronal",
            ], [total, p, d, i, id_, nf, ok, nok, n_patron]):
                col_m.metric(lbl, val)

            if "Confiança" in df.columns:
                n_manual = len(df[df["Confiança"] == "manual"])
                n_alta   = len(df[df["Confiança"] == "alta"])
                n_media  = len(df[df["Confiança"] == "media"])
                n_baixa  = len(df[df["Confiança"] == "baixa"])
                partes = []
                if n_manual: partes.append(f"🔧 {n_manual} manual")
                if n_alta:   partes.append(f"🟢 {n_alta} alta")
                if n_media:  partes.append(f"🟡 {n_media} média")
                if n_baixa:  partes.append(f"🔴 {n_baixa} baixa")
                if partes:
                    st.info(f"🤖 Classificação: {' · '.join(partes)}")

            def hl(row):
                t  = str(row.get("Tipo", ""))
                tf = str(row.get("Tipo Folha", ""))
                g  = str(row.get("Grupo", ""))
                if tf == "Empresa" or g == "Encargo Patronal":
                    return ["background-color:#E8D5FF"] * len(row)
                if t == "Provento":      return ["background-color:#d4edda"] * len(row)
                if t == "Desconto":      return ["background-color:#f8d7da"] * len(row)
                if t == "Informativa":   return ["background-color:#cce5ff"] * len(row)
                if t == "Inf. Dedutora": return ["background-color:#fff3cd"] * len(row)
                return ["background-color:#e2e3e5"] * len(row)

            st.dataframe(
                df.head(150).style.apply(hl, axis=1),
                use_container_width=True,
            )

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════
    # ETAPA 2
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("## 📥 Etapa 2 — Importar Excel Preenchido → Gerar Arquivos Finais")
    st.markdown(
        "1. Baixe o Excel da Etapa 1 · "
        "2. Revise/ajuste Conta Débito e Conta Crédito se necessário · "
        "3. Faça upload e clique em **▶ Gerar**"
    )

    excel_preenchido = st.file_uploader("4️⃣ Excel Preenchido (.xlsx)", type=["xlsx","xls"], key="excel_etapa2")
    col_btn3, _ = st.columns(2)
    with col_btn3:
        gerar_finais = st.button(
            "▶ Gerar Arquivos Finais",
            disabled=(excel_preenchido is None),
            use_container_width=True,
            type="primary",
        )

    if gerar_finais and excel_preenchido:
        log = list(st.session_state.log)
        log.append("[Etapa 2] Iniciando...")
        with st.spinner("Lendo Excel preenchido..."):
            df_preen = ler_excel_preenchido(excel_preenchido.read(), log)
        if df_preen is not None:
            with st.spinner("Gerando arquivos finais..."):
                evento_bytes, integra_bytes = gerar_arquivos_finais(df_preen, cod_empresa, log)
            st.session_state.evento_xlsx = evento_bytes
            st.session_state.integra_xls = integra_bytes
        st.session_state.log = log
        st.rerun()

    if st.session_state.evento_xlsx is not None:
        st.success("✅ Arquivos finais gerados!")
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.download_button(
                label="⬇ Baixar evento exemplo.xlsx",
                data=st.session_state.evento_xlsx,
                file_name="evento exemplo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        with col_d2:
            st.download_button(
                label="⬇ Baixar integra exemplo.xlsx",
                data=st.session_state.integra_xls,
                file_name="integra exemplo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

    # ── Log ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**Log de processamento**")
    log_texto = "\n".join(st.session_state.log)
    tem_erro  = any(str(l).upper().startswith("ERRO") for l in st.session_state.log)
    cor_borda = "#D32F2F" if tem_erro else "#388E3C"
    st.markdown(
        f"""
        <div style="background:#FCFCFC; border:1px solid {cor_borda};
                    border-radius:6px; padding:14px;
                    font-family:Consolas,monospace; font-size:13px;
                    white-space:pre-wrap; max-height:300px;
                    overflow-y:auto; color:#1F1F1F;">{log_texto}</div>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
